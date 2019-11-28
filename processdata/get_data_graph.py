from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment
import matplotlib
import matplotlib.pyplot as plt
from GlobalVar import Logger, Profile, GloVar
# 设置中文字体和负号正常显示
matplotlib.rcParams['font.sans-serif'] = ['SimHei']
matplotlib.rcParams['axes.unicode_minus'] = False


class GenerateDataGraph:
    def __init__(self, file):
        self.file = file
        # self.standard_value = int(Profile(type='read', file=GloVar.config_file_path, section='standard', option='critical_value').value)
        self.standard_value = 100

    # 获取某个单元格的值
    def get_cell_value(self, sheet, row, column):
        cell_value = sheet.cell(row=row, column=column).value
        return cell_value

    # 获取某行所有值
    def get_row_values(self, sheet, row):
        columns = sheet.max_column
        row_data = []
        for i in range(1, columns + 1):
            cell_value = sheet.cell(row=row, column=i).value
            row_data.append(cell_value)
        # 去掉行头(在这里就代表case名)
        row_data.pop(0)
        return row_data

    # 获取某列的所有值
    def get_column_values(self, sheet, column):
        rows = sheet.max_row
        column_data = []
        for i in range(1, rows + 1):
            cell_value = sheet.cell(row=i, column=column).value
            column_data.append(cell_value)
        # 去掉列头(在这里就代表差帧)
        column_data.pop(0)
        return column_data

    # 从excel表中获取数据(每一个sheet表对应一个数据图)
    def get_data(self):
        # 载入excel
        work_book = load_workbook(self.file)
        # 获取所有sheet
        sheet_list = work_book.sheetnames
        # 设置data_list存放获取到的数据([[cases, frame_gap], [cases, frame_gap]])
        data_list = []
        for sheet in sheet_list:
            case_type = sheet
            work_sheet = work_book[sheet]
            cases = self.get_column_values(sheet=work_sheet, column=1)
            frame_gap = self.get_column_values(sheet=work_sheet, column=6)
            data_list.append([case_type, cases, frame_gap])
        work_book.close()
        return data_list


    def drawing(self, file_name, case_list, frame_gap_list):
        # 水平条形图
        """
        绘制水平条形图方法barh
        参数一：y轴
        参数二：x轴
        """
        # 根据数据量设置画布大小, 以及刻度范围
        case_num = len(case_list)
        # 获取标准数据(由此判断pass/failed)
        standard_list = [self.standard_value] * case_num
        # 画布大小根据case数量来
        plt.figure(figsize=(12, case_num + 2))
        # 根据在标准值和测试值中取最大值再加上20, 作为表的宽度
        chat_width = max(frame_gap_list + standard_list) + 20
        # 设置x & y 轴范围
        plt.axis([10, chat_width, -1, case_num + 1])
        # 画条形(真实数据, 横向展示)
        y_list = [num + 0.2 for num in range(case_num)]
        plt.barh(y=y_list, width=frame_gap_list, label='测试帧数', height=0.4, color='purple', alpha=0.5)
        # 画条形(标准数据, 紧挨着真实数据)
        y_list = [num - 0.2 for num in range(case_num)]
        plt.barh(y=y_list, width=standard_list, label='标准帧数', height=0.4, color='steelblue', alpha=0.7)
        # 显示case名
        plt.yticks(range(case_num), case_list)
        # 找出真实帧数和标准帧数中较大的值(组成一个列表)
        larger_value_list = []
        for i in range(case_num):
            larger_value = standard_list[i] if standard_list[i] >= frame_gap_list[i] else frame_gap_list[i]
            larger_value_list.append(larger_value)
        # 柱形图标注
        # 标注真实测量数据
        for x, y in enumerate(frame_gap_list):
            plt.text(y + 1.5, x + 0.1, '%s' % y, ha='center', va='bottom')
        # 标注标准数据
        for x, y in enumerate(standard_list):
            plt.text(y - 1.5, x - 0.3, '%s' % y, ha='center', va='bottom')
        # 标注pass/failed
        for num in range(case_num):
            if frame_gap_list[num] > standard_list[num]:
                plt.text(chat_width - 6, num - 0.1, 'failed', size=16, color='red', ha='center', va='bottom')
            else:
                plt.text(chat_width - 6, num - 0.1, 'pass', size=16, color='green', ha='center', va='bottom')
        # 获取图片title
        title = file_name.split('/')[-1].split('.')[0]
        plt.title('[' + title + ']' + '-->流畅度测试')
        plt.xlabel('-*- 间隔帧数/fps -*-')
        # 显示图例
        plt.legend(loc='upper left')
        # 避免出现title或者纵坐标显示不完整的问题
        plt.tight_layout()
        plt.savefig(file_name)
        Logger('生成数据图: ' + file_name)

    # 获取图片
    def get_graphs(self):
        data_list = self.get_data()
        for data in data_list:
            file_name = '/'.join(self.file.split('/')[:-1]) + '/' + data[0] + '.png'
            self.drawing(file_name=file_name, case_list=data[1], frame_gap_list=data[2])



if __name__=='__main__':
    file = 'D:/Code/robot/report/2019-11-27/report.xlsx'
    data_graph = GenerateDataGraph(file=file)
    data_graph.get_graphs()
