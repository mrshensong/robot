import os
import cv2
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, colors
from processdata.get_data_graph import GenerateDataGraph
from processdata.get_report import GenerateReport
from GlobalVar import MergePath, Logger, GloVar, Profile, WindowStatus


class GetStartupTime:

    def __init__(self, video_path):
        # 视频路径
        self.video_path = video_path
        # 报告路径
        self.report_path = video_path.replace('/video/', '/report/')
        if os.path.exists(self.report_path) is False:
            os.makedirs(self.report_path)
        # excel存储数据
        self.report_excel = self.report_path + '/' + 'report.xlsx'
        # 模板对比的目标图片比模板稍大一点(默认大10像素)
        self.template_edge = 10
        # 模板匹配率(匹配率大于此, 说明可以开始检测稳定帧)
        self.template_start_frame_match_threshold = 0.93
        # 稳定帧匹配率(匹配率连续大于此匹配率, 才说明帧稳定)
        self.template_stability_frame_match_threshold = 0.98
        # 基准图片和视频帧匹配率连续超过某一匹配率的次数(可以说明帧稳定)
        self.frame_stability_times = 10
        # 待处理的视频列表
        self.videos_list = []

    def match_template(self, source_img, target_img):
        """
        :param source_img: 源图像(大图)
        :param target_img: 靶子图像(小图)
        :return:
        """
        if type(source_img) is str:
            source_img = cv2.imdecode(np.fromfile(source_img, dtype=np.uint8), -1)
        if type(target_img) is str:
            target_img = cv2.imdecode(np.fromfile(target_img, dtype=np.uint8), -1)
        # 匹配方法
        match_method = cv2.TM_CCOEFF_NORMED
        # 模板匹配
        match_result = cv2.matchTemplate(source_img, target_img, match_method)
        # 查找匹配度和坐标位置
        min_threshold, max_threshold, min_threshold_position, max_threshold_position = cv2.minMaxLoc(match_result)
        # 返回最大匹配率
        return max_threshold

    # 获取roi模板的位置信息(roi图片/column行数,0-3)
    def get_position_info_from_roi(self, roi, column):
        thousands = roi[0][column] * 1000
        hundred = roi[1][column] * 100
        ten = roi[2][column] * 10
        single = roi[3][column] * 1
        num = thousands + hundred + ten + single
        return num

    # 计算开始和结束位置
    def get_start_and_end_match_threshold(self, end_mask, video_file):
        """
        获取起止点的匹配率列表
        :param end_mask: 稳定点模板
        :param video_file: 视频文件
        :return: 开始点匹配率列表, 稳定点匹配率列表
        """
        # 获取视频对象
        video_cap = cv2.VideoCapture(video_file)
        # 帧编号
        frame_id = 0
        # 起点结果是否找到标志位
        start_point_exist_flag = False
        # 模板出现标志
        template_appear_flag = False
        # 上一帧图片(前后两帧需要对比)
        last_picture = None
        # 匹配率连续大于0.99的次数(只要中断就从零计算)/原理:连续十帧匹配率大于0.99就可以认为此帧为稳定帧
        stability_num = 0
        # 上一循环是否匹配率大于0.99的标志
        cycle_flag = False
        # 起始帧和结束帧结果(frame_id, match_rate), (frame_id, match_rate)
        start_point_result = (1, 1.0)
        end_point_result = (1, 0.0)
        # 计算待检测模板在整帧中的位置(需要截取出来)
        # 获取视频的尺寸
        frame_height = video_cap.get(cv2.CAP_PROP_FRAME_HEIGHT)
        frame_width = video_cap.get(cv2.CAP_PROP_FRAME_WIDTH)
        # 获取模板的灰度图
        end_mask_gray = cv2.imdecode(np.fromfile(end_mask, dtype=np.uint8), cv2.IMREAD_GRAYSCALE)
        # 根据模板图片中第一行像素值, 找到模板所在当前帧中的位置(将带查找图片选择的稍微比模板图片大一点点每个边大20像素)
        # 行起点
        row_start = self.get_position_info_from_roi(end_mask_gray, 0) - self.template_edge
        row_start = 0 if row_start < 0 else row_start
        # 行终点
        row_end = self.get_position_info_from_roi(end_mask_gray, 1) + self.template_edge
        row_end = frame_height - 1 if row_end >= frame_height else row_end
        # 列起点
        column_start = self.get_position_info_from_roi(end_mask_gray, 2) - self.template_edge
        column_start = 0 if column_start < 0 else column_start
        # 列终点
        column_end = self.get_position_info_from_roi(end_mask_gray, 3) + self.template_edge
        column_end = frame_width - 1 if column_end >= frame_width else column_end
        # 视频读取
        while video_cap.isOpened():
            # 帧从1开始
            frame_id += 1
            successfully_read, frame = video_cap.read()
            if successfully_read is True:
                # 灰度化
                frame_gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
                # 目标图片(待匹配)
                target = frame_gray[row_start: row_end, column_start: column_end]
                # 起点
                if start_point_exist_flag is False:
                    # 如果图像第一行像素为白色(排除前十帧可能出现的乱帧)
                    if frame[0].mean() > 245 and frame_id > 10:
                        start_point_result = (frame_id, 1.0)
                        start_point_exist_flag = True
                    # 先寻找起点(没有找到起点进行下一循环)
                    continue
                # 起点之后寻找模板出现的帧
                elif template_appear_flag is False:
                    threshold = self.match_template(target, end_mask_gray)
                    if threshold >= self.template_start_frame_match_threshold:
                        template_appear_flag = True
                        last_picture = target
                    else:
                        continue
                # 终点(稳定点)
                if template_appear_flag is True:
                    match_rate = self.match_template(target, last_picture)
                    if cycle_flag is True:
                        if match_rate > self.template_stability_frame_match_threshold:
                            cycle_flag = True
                            stability_num += 1
                            if stability_num == self.frame_stability_times:
                                end_point_result = (frame_id - self.frame_stability_times, self.template_stability_frame_match_threshold)
                                break
                        else:
                            cycle_flag = False
                            stability_num = 0
                            last_picture = target
                    else:
                        if match_rate > self.template_stability_frame_match_threshold:
                            cycle_flag = True
                            stability_num = 1
                        else:
                            cycle_flag = False
                            stability_num = 0
                            last_picture = target
            else:
                break
        video_cap.release()
        return start_point_result, end_point_result

    # 处理视频, 从视频中获取处理信息
    def process_video(self, video_id, file, end_mask):
        Logger('正在计算<%s>起止点...' % file)
        # 计算起始帧和停止帧
        (start_frame, start_threshold), (end_frame, end_threshold) = self.get_start_and_end_match_threshold(end_mask=end_mask, video_file=file)
        start_frame = start_frame
        end_frame = end_frame
        frame_gap = end_frame - start_frame
        time_gap = int(frame_gap * (1000 / 120))
        Logger('%s-->起始点: 帧> %d, 匹配率> %.4f' % (file, start_frame, start_threshold))
        Logger('%s-->终止点: 帧> %d, 匹配率> %.4f' % (file, end_frame, end_threshold))
        return {'次序':video_id, '开始帧':start_frame, '开始帧匹配率':start_threshold, '结束帧':end_frame, '结束帧匹配率':end_threshold, '差帧':frame_gap, '耗时':time_gap}

    # 处理一条case(可能含有多次执行产生的多个视频, 传入的参数为产生的这些视频的当前目录路径)
    def process_case(self, video_name_path):
        video_count = 0
        video_valid_count = 0
        video_info_list = []
        video_files = os.listdir(video_name_path)
        case_name = video_name_path.split('/')[-1]
        video_name_path_cut_list = video_name_path.split('/')
        new_video_name_path_cut_list = video_name_path_cut_list[:-4] + ['template'] + video_name_path_cut_list[-2:]
        end_mask = '/'.join(new_video_name_path_cut_list).replace('/video/', '/picture/') + '.bmp'
        # 通过读取配置文件获取标准耗时(单位/ms)
        option = '/'.join(video_name_path_cut_list[-2:])
        # standard_time_gap = 800
        standard_time_gap = int(Profile(type='read', file=GloVar.config_file_path, section='standard', option=option).value)
        # 标准差帧
        standard_frame_gap = int(standard_time_gap / (1000 / 120))
        # 差帧总和 & 平均差帧
        sum_frame_gap, average_frame_gap = 0, 0
        # 处理一个case中的多次执行产生的视频
        for video_file in video_files:
            (file_text, extension) = os.path.splitext(video_file)
            if extension in ['.mp4', '.MP4', '.avi', '.AVI']:
                video_count += 1
                file = MergePath([video_name_path, video_file]).merged_path
                # 获取到视频处理后的信息
                video_info = self.process_video(video_id=video_count, file=file, end_mask=end_mask)
                # video_info_list.append(video_info)
                # 根据视频名的数字插入info(确保视频的先后顺序和info对应)
                video_info_list.insert(int(file_text), video_info)
                if video_info['开始帧'] != 1 and video_info['结束帧'] != 1 and video_info['开始帧'] != video_info['结束帧']:
                    video_valid_count += 1
                    sum_frame_gap += int(video_info['差帧'])
        # 避免分母为0的情况
        if sum_frame_gap == 0:
            video_valid_count = 1
        average_frame_gap = int(sum_frame_gap / video_valid_count)
        average_time_gap = int((sum_frame_gap * (1000 / 120)) / video_valid_count)
        # 获取状态
        if average_frame_gap == 0 or video_valid_count < video_count:
            status = 'error'
        else:
            if average_time_gap > standard_time_gap:
                status = 'failed'
            else:
                status = 'pass'
        # 当前case中有某次计算出异常的问题, 判断是否需要重新测试
        retest_flag = 'YES' if status == 'error' else 'NO'
        # 返回的信息跟excel模型相似(列表第一个参数代表case类型, 也就是sheet名字)
        return [case_name, video_count, video_info_list, standard_frame_gap, standard_time_gap, average_frame_gap, average_time_gap, status, retest_flag]

    # 获取起止点以及对应匹配率等等
    def get_all_video_start_and_end_points(self):
        """
        此处传入视频路径(算出路径下所有case视频的起止点)
        :return:
        """
        # 获取视频名文件夹(如:D:/Code/robot/video/2019-11-27/15-10-55/测试/点击设置, 里面装有1/2/3/4/5.MP4这样的记录次数的文件)
        video_name_path_list = []
        for home, dirs, files in os.walk(self.video_path):
            if len(files) > 0:
                # 修正路径
                home = MergePath([home]).merged_path
                video_name_path_list.append(home)
        # 用dict保存每一个case类型(相当于excel的sheet名)
        case_type_dict = {}
        # 遍历每一个case视频(这样的路径: D:/Code/robot/video/2019-11-27/测试/点击设置)
        for video_name_path in video_name_path_list:
            case_type = video_name_path.split('/')[-2:-1][0]
            case_data = self.process_case(video_name_path=video_name_path)
            if case_type not in list(case_type_dict.keys()):
                case_type_dict[case_type] = []
            case_type_dict[case_type].append(case_data)
        # 返回(如:{'滑动':[[桌面滑动], [设置滑动]], '点击':[[点击设置], [点击地图]]})
        return case_type_dict

    # 将得到的数据写入excel
    def write_data_to_excel(self, file, original_data_dict):
        # 开始在excel中写入数据
        work_book = Workbook()
        # 样式/居中
        align = Alignment(horizontal='center', vertical='center')
        # 单元格填充颜色
        yellow_background = PatternFill("solid", fgColor="FFFF66")
        green_background = PatternFill("solid", fgColor="0099CC")
        gray_background = PatternFill("solid", fgColor="999999")
        # 直接在original_data_dict取出数据即可
        for key, data in original_data_dict.items():
            # 创建sheet表
            work_sheet = work_book.create_sheet(key)
            # 创建表头
            head = ['用例', '次数', '次序', '开始帧', '开始帧匹配率', '结束帧', '结束帧匹配率', '差帧', '耗时', '标准差帧', '标准耗时', '平均差帧', '平均耗时', '状态', '重新测试']
            work_sheet.append(head)
            # 表头背景颜色
            for i in range(1, len(head)+1):
                work_sheet.cell(1, i).alignment = align
                work_sheet.cell(1, i).fill = gray_background
            # 当前可用行号(第二行可用)
            current_row = 2
            # 背景颜色根据此数来确定(奇数为yellow, 偶数为green)
            background_color_select_num = 1
            for i in range(len(data)):
                background_color = green_background if background_color_select_num % 2 else yellow_background
                # 写一条case的数据(返回可用的行号)
                current_row = self.write_case_to_excel(work_sheet, current_row, data[i], align, background_color)
                background_color_select_num += 1
        # 删除掉第一个空白sheet
        work_book.remove(work_book['Sheet'])
        Logger('生成数据表: ' + file)
        work_book.save(file)
        work_book.close()

    # 将case写入sheet表中
    def write_case_to_excel(self, sheet, current_row, case_data, align, background_color):
        # 占用行直接获取次数即可(list第二个数为次数, 直接拿来用)
        occupied_row = int(case_data[1])
        # 下一个case当前行(也就是写完这个case后的下一行)
        next_current_row = current_row + occupied_row
        # 如果执行了多次
        if occupied_row > 1:
            # 合并用例
            sheet.merge_cells('A'+str(current_row) + ':' + 'A' +str(next_current_row -1))
            # case名
            sheet.cell(current_row, 1).value = case_data[0]
            sheet.cell(current_row, 1).alignment = align
            sheet.cell(current_row, 1).fill = background_color
            # 合并次数
            sheet.merge_cells('B' + str(current_row) + ':' + 'B' + str(next_current_row - 1))
            # 次数
            sheet.cell(current_row, 2).value = case_data[1]
            sheet.cell(current_row, 2).alignment = align
            sheet.cell(current_row, 2).fill = background_color
            # 装入每一次测试数据(从一个case产生多个视频)
            for id, data in enumerate(case_data[2]):
                sheet.cell(current_row + id, 3).value = data['次序']
                sheet.cell(current_row + id, 3).alignment = align
                sheet.cell(current_row + id, 3).fill = background_color
                sheet.cell(current_row + id, 4).value = data['开始帧']
                sheet.cell(current_row + id, 4).alignment = align
                sheet.cell(current_row + id, 4).fill = background_color
                sheet.cell(current_row + id, 5).value = data['开始帧匹配率']
                sheet.cell(current_row + id, 5).alignment = align
                sheet.cell(current_row + id, 5).fill = background_color
                sheet.cell(current_row + id, 6).value = data['结束帧']
                sheet.cell(current_row + id, 6).alignment = align
                sheet.cell(current_row + id, 6).fill = background_color
                sheet.cell(current_row + id, 7).value = data['结束帧匹配率']
                sheet.cell(current_row + id, 7).alignment = align
                sheet.cell(current_row + id, 7).fill = background_color
                sheet.cell(current_row + id, 8).value = data['差帧']
                sheet.cell(current_row + id, 8).alignment = align
                sheet.cell(current_row + id, 8).fill = background_color
                sheet.cell(current_row + id, 9).value = data['耗时']
                sheet.cell(current_row + id, 9).alignment = align
                sheet.cell(current_row + id, 9).fill = background_color
            # 合并标准差帧
            sheet.merge_cells('J' + str(current_row) + ':' + 'J' + str(next_current_row - 1))
            sheet.cell(current_row, 10).value = case_data[3]
            sheet.cell(current_row, 10).alignment = align
            sheet.cell(current_row, 10).fill = background_color
            # 合并标准耗时
            sheet.merge_cells('K' + str(current_row) + ':' + 'K' + str(next_current_row - 1))
            sheet.cell(current_row, 11).value = case_data[4]
            sheet.cell(current_row, 11).alignment = align
            sheet.cell(current_row, 11).fill = background_color
            # 合并平均差帧
            sheet.merge_cells('L' + str(current_row) + ':' + 'L' + str(next_current_row - 1))
            sheet.cell(current_row, 12).value = case_data[5]
            sheet.cell(current_row, 12).alignment = align
            sheet.cell(current_row, 12).fill = background_color
            # 合并平均耗时
            sheet.merge_cells('M' + str(current_row) + ':' + 'M' + str(next_current_row - 1))
            sheet.cell(current_row, 13).value = case_data[6]
            sheet.cell(current_row, 13).alignment = align
            sheet.cell(current_row, 13).fill = background_color
            # 合并状态
            sheet.merge_cells('N' + str(current_row) + ':' + 'N' + str(next_current_row - 1))
            sheet.cell(current_row, 14).value = case_data[7]
            sheet.cell(current_row, 14).alignment = align
            if case_data[7] == 'failed':
                background_color = PatternFill("solid", fgColor=colors.RED)
            elif case_data[7] == 'error':
                background_color = PatternFill("solid", fgColor=colors.BLUE)
            elif case_data[7] == 'pass':
                background_color = PatternFill("solid", fgColor=colors.GREEN)
            sheet.cell(current_row, 14).fill = background_color
            # 是否需要重新测试
            sheet.merge_cells('O' + str(current_row) + ':' + 'O' + str(next_current_row - 1))
            sheet.cell(current_row, 15).value = case_data[8]
            sheet.cell(current_row, 15).alignment = align
            sheet.cell(current_row, 13).fill = background_color
        # 只有执行了一次的数据
        else:
            sheet.cell(current_row, 1).value = case_data[0]
            sheet.cell(current_row, 1).alignment = align
            sheet.cell(current_row, 1).fill = background_color
            sheet.cell(current_row, 2).value = case_data[1]
            sheet.cell(current_row, 2).alignment = align
            sheet.cell(current_row, 2).fill = background_color
            sheet.cell(current_row, 3).value = case_data[2][0]['次序']
            sheet.cell(current_row, 3).alignment = align
            sheet.cell(current_row, 3).fill = background_color
            sheet.cell(current_row, 4).value = case_data[2][0]['开始帧']
            sheet.cell(current_row, 4).alignment = align
            sheet.cell(current_row, 4).fill = background_color
            sheet.cell(current_row, 5).value = case_data[2][0]['开始帧匹配率']
            sheet.cell(current_row, 5).alignment = align
            sheet.cell(current_row, 5).fill = background_color
            sheet.cell(current_row, 6).value = case_data[2][0]['结束帧']
            sheet.cell(current_row, 6).alignment = align
            sheet.cell(current_row, 6).fill = background_color
            sheet.cell(current_row, 7).value = case_data[2][0]['结束帧匹配率']
            sheet.cell(current_row, 7).alignment = align
            sheet.cell(current_row, 7).fill = background_color
            sheet.cell(current_row, 8).value = case_data[2][0]['差帧']
            sheet.cell(current_row, 8).alignment = align
            sheet.cell(current_row, 8).fill = background_color
            sheet.cell(current_row, 9).value = case_data[2][0]['耗时']
            sheet.cell(current_row, 9).alignment = align
            sheet.cell(current_row, 9).fill = background_color
            sheet.cell(current_row, 10).value = case_data[3]
            sheet.cell(current_row, 10).alignment = align
            sheet.cell(current_row, 10).fill = background_color
            sheet.cell(current_row, 11).value = case_data[4]
            sheet.cell(current_row, 11).alignment = align
            sheet.cell(current_row, 11).fill = background_color
            sheet.cell(current_row, 12).value = case_data[5]
            sheet.cell(current_row, 12).alignment = align
            sheet.cell(current_row, 12).fill = background_color
            sheet.cell(current_row, 13).value = case_data[6]
            sheet.cell(current_row, 13).alignment = align
            sheet.cell(current_row, 13).fill = background_color
            sheet.cell(current_row, 14).value = case_data[7]
            sheet.cell(current_row, 14).alignment = align
            if case_data[7] == 'failed':
                background_color = PatternFill("solid", fgColor=colors.RED)
            elif case_data[7] == 'error':
                background_color = PatternFill("solid", fgColor=colors.BLUE)
            elif case_data[7] == 'pass':
                background_color = PatternFill("solid", fgColor=colors.GREEN)
            sheet.cell(current_row, 14).fill = background_color
            sheet.cell(current_row, 15).value = case_data[8]
            sheet.cell(current_row, 15).alignment = align
            sheet.cell(current_row, 15).fill = background_color
        return next_current_row

    # 通过excel获取柱形图
    def get_graph_data(self, graph_path, case_date_dict):
        generate_graph = GenerateDataGraph(graph_path=graph_path, data_dict=case_date_dict)
        generate_graph.get_graphs()

    # 生成html并保存
    def get_report(self, report_path, case_data_dict):
        generate_report = GenerateReport(report_path=report_path, data_dict=case_data_dict)
        generate_report.save_html()

    # 结合(1.生成excel, 2.通过excel获取柱形图, 3.生成html并保存)
    def data_processing(self):
        # 获取数据(dict类型)
        case_data_dict = self.get_all_video_start_and_end_points()
        # 保存excel
        self.write_data_to_excel(file=self.report_excel, original_data_dict=case_data_dict)
        # # 通过excel获取柱形图
        self.get_graph_data(graph_path=self.report_path, case_date_dict=case_data_dict)
        # # 生成html并保存
        self.get_report(report_path=self.report_path, case_data_dict=case_data_dict)
        # 将数据处理完成标志位置位
        GloVar.data_process_finished_flag = True
        Logger('data process finished!')
        WindowStatus.operating_status = '空闲状态/测试结束'


if __name__=='__main__':
    # video_path = 'D:/Code/robot/video/2019-10-15'
    # video_path = 'D:/Code/robot/video/2019-12-03/点击/点击设置'
    video_path = 'D:/Code/robot_exe/video/2020-03-23/15-13-51/启动/启动音乐'
    test = GetStartupTime(video_path=video_path)
    test.data_processing()