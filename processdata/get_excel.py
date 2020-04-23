from GlobalVar import *
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, colors


class GenerateExcel:
    def __init__(self, excel_path, data_dict):
        self.excel_path = excel_path
        self.data_dict = data_dict
        self.excel_file = '/'.join([excel_path, 'report.xlsx'])
        # 样式/居中
        self.align = Alignment(horizontal='center', vertical='center')
        # 单元格填充颜色
        self.yellow_background = PatternFill("solid", fgColor="FFFF66")
        self.green_background = PatternFill("solid", fgColor="0099CC")
        self.gray_background = PatternFill("solid", fgColor="999999")

    # 将得到的数据写入excel
    def write_data_to_excel(self):
        # 开始在excel中写入数据
        work_book = Workbook()
        # 直接在original_data_dict取出数据即可
        for key, data in self.data_dict.items():
            # 创建sheet表
            work_sheet = work_book.create_sheet(key)
            # 创建表头
            head = ['用例', '次数', '次序', '时间', '开始帧', '开始帧匹配率', '结束帧', '结束帧匹配率', '差帧', '耗时', '标准差帧',
                    '标准耗时', '平均差帧', '平均耗时', '状态', '重新测试']
            work_sheet.append(head)
            # 表头背景颜色
            for i in range(1, len(head) + 1):
                work_sheet.cell(1, i).alignment = self.align
                work_sheet.cell(1, i).fill = self.gray_background
            # 当前可用行号(第二行可用)
            current_row = 2
            # 背景颜色根据此数来确定(奇数为yellow, 偶数为green)
            background_color_select_num = 1
            for i in range(len(data)):
                background_color = self.green_background if background_color_select_num % 2 else self.yellow_background
                # 写一条case的数据(返回可用的行号)
                current_row = self.write_case_to_excel(work_sheet, current_row, data[i], self.align, background_color)
                background_color_select_num += 1
        # 删除掉第一个空白sheet
        work_book.remove(work_book['Sheet'])
        Logger('生成数据表: ' + self.excel_file)
        work_book.save(self.excel_file)
        work_book.close()


    # 将case写入sheet表中
    @staticmethod
    def write_case_to_excel(sheet, current_row, case_data, align, background_color):
        # 占用行直接获取次数即可(list第二个数为次数, 直接拿来用)
        occupied_row = int(case_data[1])
        # 下一个case当前行(也就是写完这个case后的下一行)
        next_current_row = current_row + occupied_row
        # 如果执行了多次
        if occupied_row > 1:
            # 合并用例
            sheet.merge_cells('A' + str(current_row) + ':' + 'A' + str(next_current_row - 1))
            # case名
            sheet.cell(current_row, 1).value = case_data[0]
            sheet.cell(current_row, 1).alignment = align
            sheet.cell(current_row, 1).fill = background_color
            # 合并次数
            sheet.merge_cells('B' + str(current_row) + ':' + 'B' + str(next_current_row - 1))
            # 次数
            sheet.cell(current_row, 2).value = case_data[1]
            sheet.cell(current_row, 2).alignment = align
            sheet.cell(current_row, 2).fill = background_color
            # 装入每一次测试数据(从一个case产生多个视频)
            for id, data in enumerate(case_data[2]):
                sheet.cell(current_row + id, 3).value = data['次序']
                sheet.cell(current_row + id, 3).alignment = align
                sheet.cell(current_row + id, 3).fill = background_color
                sheet.cell(current_row + id, 4).value = data['时间']
                sheet.cell(current_row + id, 4).alignment = align
                sheet.cell(current_row + id, 4).fill = background_color
                sheet.cell(current_row + id, 5).value = data['开始帧']
                sheet.cell(current_row + id, 5).alignment = align
                sheet.cell(current_row + id, 5).fill = background_color
                sheet.cell(current_row + id, 6).value = data['开始帧匹配率']
                sheet.cell(current_row + id, 6).alignment = align
                sheet.cell(current_row + id, 6).fill = background_color
                sheet.cell(current_row + id, 7).value = data['结束帧']
                sheet.cell(current_row + id, 7).alignment = align
                sheet.cell(current_row + id, 7).fill = background_color
                sheet.cell(current_row + id, 8).value = data['结束帧匹配率']
                sheet.cell(current_row + id, 8).alignment = align
                sheet.cell(current_row + id, 8).fill = background_color
                sheet.cell(current_row + id, 9).value = data['差帧']
                sheet.cell(current_row + id, 9).alignment = align
                sheet.cell(current_row + id, 9).fill = background_color
                sheet.cell(current_row + id, 10).value = data['耗时']
                sheet.cell(current_row + id, 10).alignment = align
                sheet.cell(current_row + id, 10).fill = background_color
            # 合并标准差帧
            sheet.merge_cells('K' + str(current_row) + ':' + 'K' + str(next_current_row - 1))
            sheet.cell(current_row, 11).value = case_data[3]
            sheet.cell(current_row, 11).alignment = align
            sheet.cell(current_row, 11).fill = background_color
            # 合并标准耗时
            sheet.merge_cells('L' + str(current_row) + ':' + 'L' + str(next_current_row - 1))
            sheet.cell(current_row, 12).value = case_data[4]
            sheet.cell(current_row, 12).alignment = align
            sheet.cell(current_row, 12).fill = background_color
            # 合并平均差帧
            sheet.merge_cells('M' + str(current_row) + ':' + 'M' + str(next_current_row - 1))
            sheet.cell(current_row, 13).value = case_data[5]
            sheet.cell(current_row, 13).alignment = align
            sheet.cell(current_row, 13).fill = background_color
            # 合并平均耗时
            sheet.merge_cells('N' + str(current_row) + ':' + 'N' + str(next_current_row - 1))
            sheet.cell(current_row, 14).value = case_data[6]
            sheet.cell(current_row, 14).alignment = align
            sheet.cell(current_row, 14).fill = background_color
            # 合并状态
            sheet.merge_cells('O' + str(current_row) + ':' + 'O' + str(next_current_row - 1))
            sheet.cell(current_row, 15).value = case_data[7]
            sheet.cell(current_row, 15).alignment = align
            if case_data[7] == 'failed':
                background_color = PatternFill("solid", fgColor=colors.RED)
            elif case_data[7] == 'error':
                background_color = PatternFill("solid", fgColor=colors.BLUE)
            elif case_data[7] == 'pass':
                background_color = PatternFill("solid", fgColor=colors.GREEN)
            sheet.cell(current_row, 15).fill = background_color
            # 是否需要重新测试
            sheet.merge_cells('P' + str(current_row) + ':' + 'P' + str(next_current_row - 1))
            sheet.cell(current_row, 16).value = case_data[8]
            sheet.cell(current_row, 16).alignment = align
            sheet.cell(current_row, 16).fill = background_color
        # 只有执行了一次的数据
        else:
            sheet.cell(current_row, 1).value = case_data[0]
            sheet.cell(current_row, 1).alignment = align
            sheet.cell(current_row, 1).fill = background_color
            sheet.cell(current_row, 2).value = case_data[1]
            sheet.cell(current_row, 2).alignment = align
            sheet.cell(current_row, 2).fill = background_color
            sheet.cell(current_row, 3).value = case_data[2][0]['次序']
            sheet.cell(current_row, 3).alignment = align
            sheet.cell(current_row, 3).fill = background_color
            sheet.cell(current_row, 4).value = case_data[2][0]['时间']
            sheet.cell(current_row, 4).alignment = align
            sheet.cell(current_row, 4).fill = background_color
            sheet.cell(current_row, 5).value = case_data[2][0]['开始帧']
            sheet.cell(current_row, 5).alignment = align
            sheet.cell(current_row, 5).fill = background_color
            sheet.cell(current_row, 6).value = case_data[2][0]['开始帧匹配率']
            sheet.cell(current_row, 6).alignment = align
            sheet.cell(current_row, 6).fill = background_color
            sheet.cell(current_row, 7).value = case_data[2][0]['结束帧']
            sheet.cell(current_row, 7).alignment = align
            sheet.cell(current_row, 7).fill = background_color
            sheet.cell(current_row, 8).value = case_data[2][0]['结束帧匹配率']
            sheet.cell(current_row, 8).alignment = align
            sheet.cell(current_row, 8).fill = background_color
            sheet.cell(current_row, 9).value = case_data[2][0]['差帧']
            sheet.cell(current_row, 9).alignment = align
            sheet.cell(current_row, 9).fill = background_color
            sheet.cell(current_row, 10).value = case_data[2][0]['耗时']
            sheet.cell(current_row, 10).alignment = align
            sheet.cell(current_row, 10).fill = background_color
            sheet.cell(current_row, 11).value = case_data[3]
            sheet.cell(current_row, 11).alignment = align
            sheet.cell(current_row, 11).fill = background_color
            sheet.cell(current_row, 12).value = case_data[4]
            sheet.cell(current_row, 12).alignment = align
            sheet.cell(current_row, 12).fill = background_color
            sheet.cell(current_row, 13).value = case_data[5]
            sheet.cell(current_row, 13).alignment = align
            sheet.cell(current_row, 13).fill = background_color
            sheet.cell(current_row, 14).value = case_data[6]
            sheet.cell(current_row, 14).alignment = align
            sheet.cell(current_row, 14).fill = background_color
            sheet.cell(current_row, 15).value = case_data[7]
            sheet.cell(current_row, 15).alignment = align
            if case_data[7] == 'failed':
                background_color = PatternFill("solid", fgColor=colors.RED)
            elif case_data[7] == 'error':
                background_color = PatternFill("solid", fgColor=colors.BLUE)
            elif case_data[7] == 'pass':
                background_color = PatternFill("solid", fgColor=colors.GREEN)
            sheet.cell(current_row, 15).fill = background_color
            sheet.cell(current_row, 16).value = case_data[8]
            sheet.cell(current_row, 16).alignment = align
            sheet.cell(current_row, 16).fill = background_color
        return next_current_row
